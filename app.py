from flask import Flask, session, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from wtforms import StringField, SelectField, DecimalField, DateField, TextAreaField, PasswordField, BooleanField, SubmitField
from wtforms.validators import DataRequired, NumberRange, Length, Optional, Email, EqualTo, ValidationError
from datetime import datetime, date
from enum import Enum
import os
import json
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import extract, func
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import NamedStyle, Font, Alignment
import tempfile

# Configuração do Flask
app = Flask(__name__)
app.config['SECRET_KEY'] = 'sua-chave-secreta-super-segura'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///resultados_financeiros.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['CLIENTES_DIR'] = os.path.join(os.path.dirname(__file__), 'clientes')

# Inicializar SQLAlchemy
db = SQLAlchemy(app)

# Configuração do Login Manager
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor, faça login para acessar esta página.'
login_manager.login_message_category = 'info'

# Modelos
class TipoConta(Enum):
    RECEITA = "receita"
    DESPESA = "despesa"
    INVESTIMENTO = "investimento"
    IMPOSTO = "imposto"
    CUSTO = "custo"
    RETIRADA = "retirada"

class Usuario(UserMixin, db.Model):
    __tablename__ = 'usuarios'
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    nome_completo = db.Column(db.String(100), nullable=False)
    ativo = db.Column(db.Boolean, default=True)
    is_admin = db.Column(db.Boolean, default=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f'<Usuario {self.username}>'

class Categoria(db.Model):
    __tablename__ = 'categorias'
    
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    nome = db.Column(db.String(100), nullable=False)
    descricao = db.Column(db.Text)
    eh_faturamento = db.Column(db.Boolean, default=False)
    ativa = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    data_inativacao = db.Column(db.DateTime, nullable=True)  # Nova coluna
    usuario_inativacao = db.Column(db.String(100), nullable=True)  # Nova coluna
    
    contas = db.relationship('Conta', backref='categoria', lazy=True)

    def pode_ser_excluida(self):
        """Verifica se a categoria pode ser excluída (não tem contas vinculadas)"""
        return self.contas.count() == 0

    def inativar(self, usuario):
        """Inativa a categoria"""
        self.ativa = False
        self.data_inativacao = datetime.utcnow()
        self.usuario_inativacao = usuario

    def __repr__(self):
        return f'<Categoria {self.nome}>'

class Conta(db.Model):
    __tablename__ = 'contas'
    
    id = db.Column(db.Integer, primary_key=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    categoria_id = db.Column(db.Integer, db.ForeignKey('categorias.id'), nullable=False)
    data = db.Column(db.Date, nullable=False)
    valor = db.Column(db.Numeric(15, 2), nullable=False)
    descricao = db.Column(db.String(255))
    ativa = db.Column(db.Boolean, default=True)  # Nova coluna
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    data_inativacao = db.Column(db.DateTime, nullable=True)  # Nova coluna
    usuario_inativacao = db.Column(db.String(100), nullable=True)  # Nova coluna

    def inativar(self, usuario):
        """Inativa a conta"""
        self.ativa = False
        self.data_inativacao = datetime.utcnow()
        self.usuario_inativacao = usuario

    def __repr__(self):
        return f'<Conta {self.valor}>'

class Empresa(db.Model):
    __tablename__ = 'empresas'
    
    id = db.Column(db.Integer, primary_key=True)
    cnpj = db.Column(db.String(18), unique=True, nullable=False)
    nome = db.Column(db.String(100), nullable=False)
    caminho_banco = db.Column(db.String(255), nullable=False)
    data_cadastro = db.Column(db.DateTime, default=datetime.utcnow)
    ativa = db.Column(db.Boolean, default=True)
    data_inativacao = db.Column(db.DateTime, nullable=True)  # Nova coluna
    usuario_inativacao = db.Column(db.String(100), nullable=True)  # Nova coluna

    def pode_ser_excluida(self):
        """Verifica se a empresa pode ser excluída (não tem dados vinculados)"""
        from sqlalchemy import and_
        categorias_count = Categoria.query.filter_by(empresa_id=self.id).count()
        contas_count = Conta.query.filter_by(empresa_id=self.id).count()
        return categorias_count == 0 and contas_count == 0

    def inativar(self, usuario):
        """Inativa a empresa"""
        self.ativa = False
        self.data_inativacao = datetime.utcnow()
        self.usuario_inativacao = usuario

    def __repr__(self):
        return f'<Empresa {self.nome}>'


class ConfiguracaoCalculo(db.Model):
    __tablename__ = 'configuracao_calculo'
    
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(50), nullable=False, unique=True)
    categorias_positivas = db.Column(db.Text)
    categorias_negativas = db.Column(db.Text)
    ativa = db.Column(db.Boolean, default=True)
    data_atualizacao = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<ConfiguracaoCalculo {self.nome}>'

# User loader
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Usuario, int(user_id))

# Context processor para empresa selecionada
@app.context_processor
def inject_empresa_selecionada():
    empresa = None
    if 'empresa_selecionada' in session:
        empresa = Empresa.query.get(session['empresa_selecionada'])
    return dict(empresa_selecionada=empresa)

# Formulários
class LoginForm(FlaskForm):
    username = StringField('Usuário', validators=[DataRequired()])
    password = PasswordField('Senha', validators=[DataRequired()])
    remember_me = BooleanField('Lembrar de mim')
    submit = SubmitField('Entrar')

class RegistroForm(FlaskForm):
    username = StringField('Nome de Usuário', validators=[DataRequired(), Length(min=3, max=80)])
    email = StringField('Email', validators=[DataRequired(), Email()])
    nome_completo = StringField('Nome Completo', validators=[DataRequired()])
    password = PasswordField('Senha', validators=[DataRequired(), Length(min=6)])
    password2 = PasswordField('Confirmar Senha', validators=[DataRequired(), EqualTo('password')])
    submit = SubmitField('Cadastrar')

    def validate_username(self, username):
        user = Usuario.query.filter_by(username=username.data).first()
        if user:
            raise ValidationError('Este nome de usuário já está em uso.')

    def validate_email(self, email):
        user = Usuario.query.filter_by(email=email.data).first()
        if user:
            raise ValidationError('Este email já está cadastrado.')

class CategoriaForm(FlaskForm):
    nome = StringField('Nome da Categoria', validators=[DataRequired(), Length(min=2, max=100)])
    descricao = TextAreaField('Descrição', validators=[Optional(), Length(max=500)])
    eh_faturamento = BooleanField('Esta é a categoria de faturamento')

class ContaForm(FlaskForm):
    categoria_id = SelectField('Categoria', coerce=int, validators=[DataRequired()])
    data = DateField('Data', validators=[DataRequired()])
    valor = DecimalField('Valor', validators=[DataRequired(), NumberRange(min=0.01)], places=2)
    descricao = StringField('Descrição', validators=[Optional(), Length(max=255)])

class EmpresaForm(FlaskForm):
    cnpj = StringField('CNPJ', validators=[DataRequired(), Length(min=18, max=18)])
    nome = StringField('Nome da Empresa', validators=[DataRequired(), Length(min=3, max=100)])

# Funções de planilha
def gerar_template_planilha():
    """Gera template de planilha para importação de contas"""
    try:
        categorias = Categoria.query.filter_by(ativa=True).all()
        
        if not categorias:
            return None, "Nenhuma categoria cadastrada"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Contas"
        
        headers = ['Categoria', 'Valor', 'Data', 'Descrição']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        nomes_categorias = [cat.nome for cat in categorias]
        
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(nomes_categorias)}"',
            allow_blank=False
        )
        dv.error = 'Categoria inválida'
        dv.errorTitle = 'Erro de Validação'
        dv.prompt = 'Selecione uma categoria da lista'
        dv.promptTitle = 'Categoria'
        
        ws.add_data_validation(dv)
        dv.add('A2:A101')
        
        valor_style = NamedStyle(name="valor_br")
        valor_style.number_format = 'R$ #,##0.00'
        
        data_style = NamedStyle(name="data_br")
        data_style.number_format = 'DD/MM/AAAA'
        
        if 'valor_br' not in wb.named_styles:
            wb.add_named_style(valor_style)
        if 'data_br' not in wb.named_styles:
            wb.add_named_style(data_style)
        
        for row in range(2, 102):
            ws.cell(row=row, column=2).style = valor_style
            ws.cell(row=row, column=3).style = data_style
        
        hoje = date.today()
        from datetime import timedelta
        
        exemplos = [
            [nomes_categorias[0] if nomes_categorias else '', 1000.00, hoje, 'Exemplo de conta 1'],
            [nomes_categorias[0] if nomes_categorias else '', 500.50, hoje + timedelta(days=1), 'Exemplo de conta 2'],
            [nomes_categorias[0] if nomes_categorias else '', 750.25, hoje + timedelta(days=2), 'Exemplo de conta 3']
        ]
        
        for row, exemplo in enumerate(exemplos, 2):
            for col, valor in enumerate(exemplo, 1):
                cell = ws.cell(row=row, column=col, value=valor)
                if col == 2:
                    cell.style = valor_style
                elif col == 3:
                    cell.style = data_style
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 35
        
        ws_instrucoes = wb.create_sheet("Instruções")
        instrucoes = [
            "INSTRUÇÕES PARA IMPORTAÇÃO DE CONTAS - ELEVALUCRO",
            "",
            "1. PREENCHIMENTO DA PLANILHA:",
            "   • Use a aba 'Contas' para inserir os dados",
            "   • Preencha todas as colunas obrigatórias",
            "   • Siga os formatos indicados abaixo",
            "",
            "2. FORMATAÇÃO DOS DADOS:",
            "   • Categoria: Selecione da lista suspensa",
            "   • Valor: Use formato brasileiro (ex: 1.500,75)",
            "   • Data: Use formato DD/MM/AAAA (ex: 15/01/2025)",
            "   • Descrição: Texto livre (opcional)",
            "",
            "3. REGRAS DE IMPORTAÇÃO:",
            "   • Apenas categorias cadastradas serão aceitas",
            "   • Linhas com categorias inválidas serão ignoradas",
            "   • Valores devem ser maiores que zero",
            "   • Datas inválidas causarão erro na linha",
            "",
            "4. CATEGORIAS DISPONÍVEIS NO SISTEMA:",
        ]
        
        for i, instrucao in enumerate(instrucoes, 1):
            cell = ws_instrucoes.cell(row=i, column=1, value=instrucao)
            if "INSTRUÇÕES" in instrucao or instrucao.endswith(":"):
                cell.font = Font(bold=True)
        
        linha_atual = len(instrucoes) + 1
        for categoria in categorias:
            cell = ws_instrucoes.cell(row=linha_atual, column=1, value=f"   • {categoria.nome}")
            if categoria.eh_faturamento:
                cell.value += " (Faturamento)"
                cell.font = Font(color="008000")
            linha_atual += 1
        
        ws_instrucoes.column_dimensions['A'].width = 60
        
        return wb, None
        
    except Exception as e:
        return None, f"Erro ao gerar template: {str(e)}"

def processar_importacao_planilha(arquivo):
    """Processa a importação de contas da planilha"""
    try:
        df = pd.read_excel(arquivo, sheet_name='Contas')
        
        categorias = Categoria.query.filter_by(ativa=True).all()
        categorias_dict = {cat.nome: cat.id for cat in categorias}
        
        contas_importadas = 0
        contas_ignoradas = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                if pd.isna(row['Categoria']) or pd.isna(row['Valor']) or pd.isna(row['Data']):
                    continue
                
                categoria_nome = str(row['Categoria']).strip()
                
                if categoria_nome not in categorias_dict:
                    contas_ignoradas += 1
                    erros.append(f"Linha {index + 2}: Categoria '{categoria_nome}' não encontrada - ignorada")
                    continue
                
                try:
                    valor_str = str(row['Valor']).strip()
                    valor_str = valor_str.replace('R$', '').replace(' ', '')
                    if ',' in valor_str and '.' in valor_str:
                        valor_str = valor_str.replace('.', '').replace(',', '.')
                    elif ',' in valor_str and '.' not in valor_str:
                        valor_str = valor_str.replace(',', '.')
                    
                    valor = float(valor_str)
                    if valor <= 0:
                        erros.append(f"Linha {index + 2}: Valor deve ser maior que zero - ignorada")
                        contas_ignoradas += 1
                        continue
                except (ValueError, TypeError):
                    erros.append(f"Linha {index + 2}: Valor inválido '{row['Valor']}' - ignorada")
                    contas_ignoradas += 1
                    continue
                
                try:
                    if isinstance(row['Data'], str):
                        data_str = row['Data'].strip()
                        if '/' in data_str:
                            data = datetime.strptime(data_str, '%d/%m/%Y').date()
                        else:
                            data = datetime.strptime(data_str, '%Y-%m-%d').date()
                    else:
                        data = row['Data'].date() if hasattr(row['Data'], 'date') else row['Data']
                except (ValueError, TypeError):
                    erros.append(f"Linha {index + 2}: Data inválida '{row['Data']}' (use DD/MM/AAAA) - ignorada")
                    contas_ignoradas += 1
                    continue
                
                conta = Conta(
                    categoria_id=categorias_dict[categoria_nome],
                    valor=valor,
                    data=data,
                    descricao=str(row['Descrição']) if not pd.isna(row['Descrição']) else None
                )
                
                db.session.add(conta)
                contas_importadas += 1
                
            except Exception as e:
                erros.append(f"Linha {index + 2}: Erro inesperado - {str(e)}")
                contas_ignoradas += 1
        
        db.session.commit()
        
        return {
            'sucesso': True,
            'contas_importadas': contas_importadas,
            'contas_ignoradas': contas_ignoradas,
            'erros': erros
        }
        
    except Exception as e:
        db.session.rollback()
        return {
            'sucesso': False,
            'erro': f"Erro ao processar planilha: {str(e)}"
        }

# Funções de cálculo
def calcular_por_configuracao_mes(config, mes, ano):
    """Calcula um indicador para um mês específico"""
    if not config:
        return 0

    valor_total = 0

    try:
        if config.categorias_positivas:
            categorias_pos = json.loads(config.categorias_positivas)
            valor_positivo = db.session.query(func.sum(Conta.valor)).filter(
                Conta.categoria_id.in_(categorias_pos),
                extract('month', Conta.data) == mes,
                extract('year', Conta.data) == ano
            ).scalar() or 0
            valor_total += float(valor_positivo)

        if config.categorias_negativas:
            categorias_neg = json.loads(config.categorias_negativas)
            valor_negativo = db.session.query(func.sum(Conta.valor)).filter(
                Conta.categoria_id.in_(categorias_neg),
                extract('month', Conta.data) == mes,
                extract('year', Conta.data) == ano
            ).scalar() or 0
            valor_total -= float(valor_negativo)

        return valor_total
        
    except Exception as e:
        print(f"Erro ao calcular configuração {config.nome}: {e}")
        return 0

def calcular_dados_grafico(ano):
    """Calcula os dados mensais para os gráficos"""
    dados = {
        'meses': ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'],
        'lucro_bruto': [0] * 12,
        'lucro_liquido': [0] * 12,
        'reserva_caixa': [0] * 12,
        'faturamento': [0] * 12
    }

    try:
        config_lucro_bruto = ConfiguracaoCalculo.query.filter_by(nome='lucro_bruto', ativa=True).first()
        config_lucro_liquido = ConfiguracaoCalculo.query.filter_by(nome='lucro_liquido', ativa=True).first()
        config_reserva_caixa = ConfiguracaoCalculo.query.filter_by(nome='reserva_caixa', ativa=True).first()
        categoria_faturamento = Categoria.query.filter_by(eh_faturamento=True).first()

        for mes in range(1, 13):
            dados['lucro_bruto'][mes-1] = calcular_por_configuracao_mes(config_lucro_bruto, mes, ano)
            dados['lucro_liquido'][mes-1] = calcular_por_configuracao_mes(config_lucro_liquido, mes, ano)
            dados['reserva_caixa'][mes-1] = calcular_por_configuracao_mes(config_reserva_caixa, mes, ano)
            
            if categoria_faturamento:
                faturamento_mes = db.session.query(func.sum(Conta.valor)).filter(
                    Conta.categoria_id == categoria_faturamento.id,
                    extract('month', Conta.data) == mes,
                    extract('year', Conta.data) == ano
                ).scalar() or 0
                dados['faturamento'][mes-1] = float(faturamento_mes)

    except Exception as e:
        print(f"Erro ao calcular dados do gráfico: {e}")

    return dados

def calcular_indicadores_anuais(ano):
    """Calcula os indicadores anuais totais"""
    try:
        faturamento = 0
        categoria_faturamento = Categoria.query.filter_by(eh_faturamento=True).first()
        if categoria_faturamento:
            faturamento_query = db.session.query(func.sum(Conta.valor)).filter(
                Conta.categoria_id == categoria_faturamento.id,
                extract('year', Conta.data) == ano
            ).scalar()
            faturamento = float(faturamento_query) if faturamento_query else 0

        config_lucro_bruto = ConfiguracaoCalculo.query.filter_by(nome='lucro_bruto', ativa=True).first()
        config_lucro_liquido = ConfiguracaoCalculo.query.filter_by(nome='lucro_liquido', ativa=True).first()
        config_reserva_caixa = ConfiguracaoCalculo.query.filter_by(nome='reserva_caixa', ativa=True).first()

        lucro_bruto = sum(calcular_por_configuracao_mes(config_lucro_bruto, mes, ano) for mes in range(1, 13))
        lucro_liquido = sum(calcular_por_configuracao_mes(config_lucro_liquido, mes, ano) for mes in range(1, 13))
        reserva_caixa = sum(calcular_por_configuracao_mes(config_reserva_caixa, mes, ano) for mes in range(1, 13))

        return {
            'faturamento': faturamento,
            'lucro_bruto': lucro_bruto,
            'lucro_liquido': lucro_liquido,
            'reserva_caixa': reserva_caixa
        }
    except Exception as e:
        print(f"Erro ao calcular indicadores anuais: {e}")
        return {
            'faturamento': 0,
            'lucro_bruto': 0,
            'lucro_liquido': 0,
            'reserva_caixa': 0
        }

# Filtros para templates
@app.template_filter('from_json')
def from_json_filter(value):
    if value:
        try:
            return json.loads(value)
        except:
            return []
    return []

@app.template_filter('moeda_br')
def moeda_br_filter(valor):
    """Formata valor para moeda brasileira"""
    try:
        if isinstance(valor, str):
            valor_limpo = valor.replace('R$', '').replace(' ', '').strip()
            if ',' in valor_limpo:
                valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
            valor_float = float(valor_limpo)
        else:
            valor_float = float(valor)
        
        return f"R$ {valor_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return "R$ 0,00"

# Rotas de Autenticação
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('empresas'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = Usuario.query.filter_by(username=form.username.data).first()
        
        if user and user.check_password(form.password.data) and user.ativo:
            login_user(user, remember=form.remember_me.data)
            next_page = request.args.get('next')
            if not next_page or not next_page.startswith('/'):
                next_page = url_for('dashboard')
            return redirect(next_page)
        else:
            flash('Usuário ou senha inválidos', 'danger')
    
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logout realizado com sucesso', 'info')
    return redirect(url_for('login'))

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    form = RegistroForm()
    if form.validate_on_submit():
        user = Usuario(
            username=form.username.data,
            email=form.email.data,
            nome_completo=form.nome_completo.data,
            is_admin=False
        )
        user.set_password(form.password.data)
        
        db.session.add(user)
        db.session.commit()
        
        flash('Usuário cadastrado com sucesso!', 'success')
        return redirect(url_for('login'))
    
    return render_template('registro.html', form=form)

# Rotas principais
@app.route('/dashboard')
@login_required
def dashboard():
    try:
        total_categorias = Categoria.query.filter_by(ativa=True).count()
        total_contas = Conta.query.count()
        
        ano_atual = datetime.now().year
        ano_anterior = ano_atual - 1
        
        anos_disponiveis_query = db.session.query(
            db.extract('year', Conta.data).label('ano')
        ).distinct().order_by(db.desc('ano')).all()
        anos_disponiveis = [str(int(ano.ano)) for ano in anos_disponiveis_query if int(ano.ano) != ano_atual]
        
        categorias_query = Categoria.query.filter_by(ativa=True).all()
        categorias_disponiveis = []
        for cat in categorias_query:
            categorias_disponiveis.append({
                'id': cat.id,
                'nome': cat.nome,
                'eh_faturamento': cat.eh_faturamento
            })
        
        indicadores = calcular_indicadores_anuais(ano_atual)
        indicadores_ano_anterior = calcular_indicadores_anuais(ano_anterior)
        
        dados_grafico = calcular_dados_grafico(ano_atual)
        dados_grafico_anterior = calcular_dados_grafico(ano_anterior)
        
    except Exception as e:
        print(f"Erro no dashboard: {e}")
        total_categorias = 0
        total_contas = 0
        anos_disponiveis = []
        categorias_disponiveis = []
        indicadores = {
            'faturamento': 0,
            'lucro_bruto': 0,
            'lucro_liquido': 0,
            'reserva_caixa': 0
        }
        indicadores_ano_anterior = indicadores.copy()
        dados_grafico = {
            'meses': ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'],
            'lucro_bruto': [0] * 12,
            'lucro_liquido': [0] * 12,
            'reserva_caixa': [0] * 12,
            'faturamento': [0] * 12
        }
        dados_grafico_anterior = dados_grafico.copy()
        ano_atual = datetime.now().year
        ano_anterior = ano_atual - 1
        flash('Sistema inicializado. Comece criando suas categorias!', 'info')
    
    return render_template('index.html', 
                         indicadores=indicadores,
                         indicadores_ano_anterior=indicadores_ano_anterior,
                         total_categorias=total_categorias,
                         total_contas=total_contas,
                         dados_grafico=dados_grafico,
                         dados_grafico_anterior=dados_grafico_anterior,
                         anos_disponiveis=anos_disponiveis,
                         categorias_disponiveis=categorias_disponiveis,
                         ano_atual=ano_atual,
                         ano_anterior=ano_anterior)

# Rotas de Categorias
@app.route('/categorias')
@login_required
def listar_categorias():
    try:
        empresa_id = session.get('empresa_selecionada')
        if empresa_id:
            # Mostrar ativas e inativas separadamente
            categorias_ativas = Categoria.query.filter_by(empresa_id=empresa_id, ativa=True).all()
            categorias_inativas = Categoria.query.filter_by(empresa_id=empresa_id, ativa=False).all()
        else:
            categorias_ativas = []
            categorias_inativas = []
            flash('Selecione uma empresa primeiro!', 'warning')
    except Exception as e:
        print(f"Erro ao listar categorias: {e}")
        categorias_ativas = []
        categorias_inativas = []
        flash('Erro ao carregar categorias!', 'error')
    
    form = CategoriaForm()
    return render_template('categorias/cadastro.html', 
                         categorias_ativas=categorias_ativas,
                         categorias_inativas=categorias_inativas,
                         form=form)

@app.route('/contas')
@login_required
def listar_contas():
    try:
        empresa_id = session.get('empresa_selecionada')
        if empresa_id:
            # Mostrar apenas contas ativas por padrão
            contas = Conta.query.filter_by(empresa_id=empresa_id, ativa=True).order_by(Conta.data.desc()).limit(100).all()
            categorias = Categoria.query.filter_by(empresa_id=empresa_id, ativa=True).all()
        else:
            contas = []
            categorias = []
            flash('Selecione uma empresa primeiro!', 'warning')
        
        anos_disponiveis = db.session.query(
            db.extract('year', Conta.data).label('ano')
        ).filter_by(empresa_id=empresa_id, ativa=True).distinct().order_by(db.desc('ano')).all()
        
        anos = [str(int(ano.ano)) for ano in anos_disponiveis]
        
    except Exception as e:
        print(f"Erro ao listar contas: {e}")
        contas = []
        categorias = []
        anos = []
        flash('Erro ao carregar contas!', 'error')
    
    form = ContaForm()
    form.categoria_id.choices = [(c.id, c.nome) for c in categorias]
    return render_template('contas/registro.html', contas=contas, categorias=categorias, form=form, anos=anos)



@app.route('/categorias/nova', methods=['POST'])
@login_required
def nova_categoria():
    form = CategoriaForm()
    if form.validate_on_submit():
        empresa_id = session.get('empresa_selecionada')
        categoria = Categoria(
            nome=form.nome.data,
            descricao=form.descricao.data,
            eh_faturamento=form.eh_faturamento.data,
            empresa_id=empresa_id
        )
        db.session.add(categoria)
        db.session.commit()
        flash('Categoria criada com sucesso!', 'success')
    return redirect(url_for('listar_categorias'))


@app.route('/categorias/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_categoria(id):
    categoria = Categoria.query.get_or_404(id)
    
    if request.method == 'GET':
        form_data = {
            'id': categoria.id,
            'nome': categoria.nome,
            'descricao': categoria.descricao,
            'eh_faturamento': categoria.eh_faturamento
        }
        return jsonify(form_data)
    
    elif request.method == 'POST':
        try:
            if request.json.get('eh_faturamento'):
                Categoria.query.filter_by(eh_faturamento=True).update({'eh_faturamento': False})
            
            categoria.nome = request.json.get('nome')
            categoria.descricao = request.json.get('descricao')
            categoria.eh_faturamento = request.json.get('eh_faturamento', False)
            
            db.session.commit()
            return jsonify({'status': 'success', 'message': 'Categoria atualizada com sucesso!'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'status': 'error', 'message': f'Erro ao atualizar categoria: {str(e)}'})

@app.route('/categorias/excluir/<int:id>')
@login_required
def excluir_categoria(id):
    try:
        categoria = Categoria.query.get_or_404(id)
        
        # Verificar se pode ser excluída ou deve ser inativada
        if categoria.pode_ser_excluida():
            # Pode excluir fisicamente
            db.session.delete(categoria)
            db.session.commit()
            flash('Categoria excluída com sucesso!', 'success')
        else:
            # Deve inativar
            categoria.inativar(current_user.username)
            db.session.commit()
            flash('Categoria inativada com sucesso! (Possui contas vinculadas)', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao processar categoria: {str(e)}', 'error')
    
    return redirect(url_for('listar_categorias'))

@app.route('/categorias/reativar/<int:id>')
@login_required
def reativar_categoria(id):
    try:
        categoria = Categoria.query.get_or_404(id)
        categoria.ativa = True
        categoria.data_inativacao = None
        categoria.usuario_inativacao = None
        db.session.commit()
        flash('Categoria reativada com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao reativar categoria: {str(e)}', 'error')
    
    return redirect(url_for('listar_categorias'))


# Rotas de Contas@app.route('/contas')

@app.route('/contas/nova', methods=['POST'])
@login_required
def nova_conta():
    form = ContaForm()
    try:
        categorias = Categoria.query.filter_by(ativa=True).all()
        form.categoria_id.choices = [(c.id, c.nome) for c in categorias]
    except Exception as e:
        flash('Crie categorias primeiro!', 'error')
        return redirect(url_for('listar_categorias'))
    
    if form.validate_on_submit():
        try:
            conta = Conta(
                categoria_id=form.categoria_id.data,
                data=form.data.data,
                valor=float(form.valor.data),
                descricao=form.descricao.data
            )
            db.session.add(conta)
            db.session.commit()
            flash('Conta registrada com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao registrar conta: {str(e)}', 'error')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{field}: {error}', 'error')
    
    return redirect(url_for('listar_contas'))

@app.route('/contas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_conta(id):
    conta = Conta.query.get_or_404(id)
    
    if request.method == 'GET':
        form_data = {
            'id': conta.id,
            'categoria_id': conta.categoria_id,
            'data': conta.data.strftime('%Y-%m-%d'),
            'valor': float(conta.valor),
            'descricao': conta.descricao
        }
        return jsonify(form_data)
    
    elif request.method == 'POST':
        try:
            conta.categoria_id = request.json.get('categoria_id')
            conta.data = datetime.strptime(request.json.get('data'), '%Y-%m-%d').date()
            conta.valor = request.json.get('valor')
            conta.descricao = request.json.get('descricao')
            
            db.session.commit()
            return jsonify({'status': 'success', 'message': 'Conta atualizada com sucesso!'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'status': 'error', 'message': f'Erro ao atualizar conta: {str(e)}'})

@app.route('/contas/excluir/<int:id>', methods=['GET', 'POST'])
@login_required
def excluir_conta(id):
    try:
        conta = Conta.query.get_or_404(id)
        
        # Contas sempre são inativadas, nunca excluídas fisicamente
        conta.inativar(current_user.username)
        db.session.commit()
        
        if request.method == 'POST':
            return jsonify({'status': 'success', 'message': 'Conta inativada com sucesso!'})
        else:
            flash('Conta inativada com sucesso!', 'success')
            return redirect(url_for('listar_contas'))
            
    except Exception as e:
        db.session.rollback()
        if request.method == 'POST':
            return jsonify({'status': 'error', 'message': f'Erro ao inativar conta: {str(e)}'})
        else:
            flash(f'Erro ao inativar conta: {str(e)}', 'error')
            return redirect(url_for('listar_contas'))

@app.route('/contas/reativar/<int:id>', methods=['POST'])
@login_required
def reativar_conta(id):
    try:
        conta = Conta.query.get_or_404(id)
        conta.ativa = True
        conta.data_inativacao = None
        conta.usuario_inativacao = None
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Conta reativada com sucesso!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'Erro ao reativar conta: {str(e)}'})


@app.route('/contas/baixar-template')
@login_required
def baixar_template_contas():
    try:
        wb, erro = gerar_template_planilha()
        
        if erro:
            flash(erro, 'warning')
            return redirect(url_for('listar_contas'))
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name='template_contas_elevalucro.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Erro ao gerar template: {str(e)}', 'error')
        return redirect(url_for('listar_contas'))

@app.route('/contas/importar', methods=['POST'])
@login_required
def importar_contas():
    try:
        if 'arquivo' not in request.files:
            flash('Nenhum arquivo selecionado', 'error')
            return redirect(url_for('listar_contas'))
        
        arquivo = request.files['arquivo']
        
        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado', 'error')
            return redirect(url_for('listar_contas'))
        
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            flash('Arquivo deve ser uma planilha Excel (.xlsx ou .xls)', 'error')
            return redirect(url_for('listar_contas'))
        
        resultado = processar_importacao_planilha(arquivo)
        
        if resultado['sucesso']:
            mensagem = f"Importação concluída! {resultado['contas_importadas']} contas importadas"
            if resultado['contas_ignoradas'] > 0:
                mensagem += f", {resultado['contas_ignoradas']} ignoradas"
            
            flash(mensagem, 'success')
            
            if resultado['erros']:
                for erro in resultado['erros'][:5]:
                    flash(erro, 'warning')
                if len(resultado['erros']) > 5:
                    flash(f"... e mais {len(resultado['erros']) - 5} erros", 'warning')
        else:
            flash(resultado['erro'], 'error')
        
    except Exception as e:
        flash(f'Erro na importação: {str(e)}', 'error')
    
    return redirect(url_for('listar_contas'))

# Rotas de Configuração
@app.route('/configuracao-calculo')
@login_required
def configuracao_calculo():
    try:
        categorias = Categoria.query.filter_by(ativa=True).all()
        configuracoes = ConfiguracaoCalculo.query.filter_by(ativa=True).all()
    except Exception as e:
        print(f"Erro ao carregar configurações: {e}")
        categorias = []
        configuracoes = []
        flash('Erro ao carregar configurações!', 'error')
    
    return render_template('configuracao/calculo.html', categorias=categorias, configuracoes=configuracoes)

@app.route('/api/configuracoes')
@login_required
def api_configuracoes():
    try:
        configuracoes = ConfiguracaoCalculo.query.filter_by(ativa=True).all()
        categorias = Categoria.query.filter_by(ativa=True).all()
        
        result = []
        for config in configuracoes:
            formula_parts = []
            
            if config.categorias_positivas:
                pos_ids = json.loads(config.categorias_positivas)
                pos_names = [cat.nome for cat in categorias if cat.id in pos_ids]
                if pos_names:
                    formula_parts.append(' + '.join(pos_names))
            
            if config.categorias_negativas:
                neg_ids = json.loads(config.categorias_negativas)
                neg_names = [cat.nome for cat in categorias if cat.id in neg_ids]
                if neg_names:
                    if formula_parts:
                        formula_parts.append(' - (' + ' + '.join(neg_names) + ')')
                    else:
                        formula_parts.append('-(' + ' + '.join(neg_names) + ')')
            
            formula = ''.join(formula_parts) if formula_parts else 'Nenhuma configuração'
            
            result.append({
                'nome': config.nome,
                'formula': formula
            })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/salvar-configuracao', methods=['POST'])
@login_required
def api_salvar_configuracao():
    try:
        data = request.get_json()
        nome = data.get('nome')
        categorias_positivas = data.get('categorias_positivas', [])
        categorias_negativas = data.get('categorias_negativas', [])
        
        config = ConfiguracaoCalculo.query.filter_by(nome=nome).first()
        if not config:
            config = ConfiguracaoCalculo(nome=nome)
        
        config.categorias_positivas = json.dumps(categorias_positivas) if categorias_positivas else None
        config.categorias_negativas = json.dumps(categorias_negativas) if categorias_negativas else None
        config.ativa = True
        
        if config.id is None:
            db.session.add(config)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Configuração salva com sucesso!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# APIs do Dashboard
@app.route('/api/dashboard-filtrado', methods=['POST'])
@login_required
def api_dashboard_filtrado():
    try:
        data = request.get_json()
        ano = data.get('ano')
        meses = data.get('meses', [])
        
        indicadores = {
            'faturamento': 0,
            'lucro_bruto': 0,
            'lucro_liquido': 0,
            'reserva_caixa': 0
        }
        
        dados_grafico = {
            'faturamento': [0] * 12,
            'lucro_bruto': [0] * 12,
            'lucro_liquido': [0] * 12,
            'reserva_caixa': [0] * 12
        }
        
        config_lucro_bruto = ConfiguracaoCalculo.query.filter_by(nome='lucro_bruto', ativa=True).first()
        config_lucro_liquido = ConfiguracaoCalculo.query.filter_by(nome='lucro_liquido', ativa=True).first()
        config_reserva_caixa = ConfiguracaoCalculo.query.filter_by(nome='reserva_caixa', ativa=True).first()
        categoria_faturamento = Categoria.query.filter_by(eh_faturamento=True).first()
        
        for mes in range(1, 13):
            dados_grafico['lucro_bruto'][mes-1] = calcular_por_configuracao_mes(config_lucro_bruto, mes, ano)
            dados_grafico['lucro_liquido'][mes-1] = calcular_por_configuracao_mes(config_lucro_liquido, mes, ano)
            dados_grafico['reserva_caixa'][mes-1] = calcular_por_configuracao_mes(config_reserva_caixa, mes, ano)
            
            if categoria_faturamento:
                faturamento_mes = db.session.query(func.sum(Conta.valor)).filter(
                    Conta.categoria_id == categoria_faturamento.id,
                    extract('month', Conta.data) == mes,
                    extract('year', Conta.data) == ano
                ).scalar() or 0
                dados_grafico['faturamento'][mes-1] = float(faturamento_mes)
            
            if mes in meses:
                indicadores['lucro_bruto'] += dados_grafico['lucro_bruto'][mes-1]
                indicadores['lucro_liquido'] += dados_grafico['lucro_liquido'][mes-1]
                indicadores['reserva_caixa'] += dados_grafico['reserva_caixa'][mes-1]
                indicadores['faturamento'] += dados_grafico['faturamento'][mes-1]
        
        return jsonify({
            'indicadores': indicadores,
            'dados_grafico': dados_grafico
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dados-categorias', methods=['POST'])
@login_required
def api_dados_categorias():
    try:
        data = request.get_json()
        categorias_ids = data.get('categorias', [])
        ano = data.get('ano')
        meses = data.get('meses', [])
        
        resultado = []
        
        for categoria_id in categorias_ids:
            categoria = Categoria.query.get(categoria_id)
            if not categoria:
                continue
            
            total = 0
            for mes in meses:
                valor_mes = db.session.query(func.sum(Conta.valor)).filter(
                    Conta.categoria_id == categoria_id,
                    extract('month', Conta.data) == mes,
                    extract('year', Conta.data) == ano
                ).scalar() or 0
                total += float(valor_mes)
            
            resultado.append({
                'id': categoria.id,
                'nome': categoria.nome,
                'valor': total
            })
        
        resultado.sort(key=lambda x: x['valor'], reverse=True)
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Rotas de Usuários
@app.route('/usuarios')
@login_required
def listar_usuarios():
    if not current_user.is_admin:
        flash('Acesso negado. Apenas administradores podem acessar esta página.', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        usuarios = Usuario.query.order_by(Usuario.data_criacao.desc()).all()
    except Exception as e:
        print(f"Erro ao listar usuários: {e}")
        usuarios = []
        flash('Erro ao carregar usuários!', 'error')
    
    return render_template('usuarios/lista.html', usuarios=usuarios)

@app.route('/usuarios/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_usuario(id):
    if not current_user.is_admin:
        return jsonify({'status': 'error', 'message': 'Acesso negado'}), 403
    
    usuario = Usuario.query.get_or_404(id)
    
    if request.method == 'GET':
        form_data = {
            'id': usuario.id,
            'nome_completo': usuario.nome_completo,
            'username': usuario.username,
            'email': usuario.email,
            'is_admin': usuario.is_admin,
            'ativo': usuario.ativo
        }
        return jsonify(form_data)
    
    elif request.method == 'POST':
        try:
            data = request.get_json()
            
            if usuario.id == current_user.id and not data.get('is_admin'):
                return jsonify({'status': 'error', 'message': 'Você não pode remover privilégios de administrador de si mesmo!'})
            
            if usuario.id == current_user.id and not data.get('ativo'):
                return jsonify({'status': 'error', 'message': 'Você não pode desativar sua própria conta!'})
            
            usuario.nome_completo = data.get('nome_completo')
            usuario.username = data.get('username')
            usuario.email = data.get('email')
            usuario.is_admin = data.get('is_admin', False)
            usuario.ativo = data.get('ativo', True)
            
            nova_senha = data.get('nova_senha')
            if nova_senha and len(nova_senha.strip()) >= 6:
                usuario.set_password(nova_senha.strip())
            
            db.session.commit()
            return jsonify({'status': 'success', 'message': 'Usuário atualizado com sucesso!'})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'status': 'error', 'message': f'Erro ao atualizar usuário: {str(e)}'})

@app.route('/usuarios/toggle/<int:id>', methods=['POST'])
@login_required
def toggle_usuario(id):
    if not current_user.is_admin:
        return jsonify({'status': 'error', 'message': 'Acesso negado'}), 403
    
    try:
        usuario = Usuario.query.get_or_404(id)
        if usuario.id == current_user.id:
            return jsonify({'status': 'error', 'message': 'Você não pode alterar o status da sua própria conta!'})
        
        usuario.ativo = not usuario.ativo
        db.session.commit()
        
        status = 'ativado' if usuario.ativo else 'desativado'
        return jsonify({'status': 'success', 'message': f'Usuário {status} com sucesso!'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'Erro ao alterar usuário: {str(e)}'})

@app.route('/usuarios/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_usuario(id):
    if not current_user.is_admin:
        return jsonify({'status': 'error', 'message': 'Acesso negado'}), 403
    
    try:
        usuario = Usuario.query.get_or_404(id)
        
        if usuario.id == current_user.id:
            return jsonify({'status': 'error', 'message': 'Você não pode excluir sua própria conta!'})
        
        if usuario.is_admin:
            outros_admins = Usuario.query.filter(Usuario.is_admin == True, Usuario.id != id, Usuario.ativo == True).count()
            if outros_admins == 0:
                return jsonify({'status': 'error', 'message': 'Não é possível excluir o último administrador ativo!'})
        
        db.session.delete(usuario)
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': 'Usuário excluído com sucesso!'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'Erro ao excluir usuário: {str(e)}'})

# Rotas de Empresas
@app.route('/empresas', methods=['GET'])
@login_required
def empresas():
    empresas = Empresa.query.filter_by(ativa=True).all()
    form = EmpresaForm()
    return render_template('empresas/cadastro_empresas.html', empresas=empresas, form=form)

@app.route('/empresas/nova', methods=['POST'])
@login_required
def nova_empresa():
    form = EmpresaForm()
    if form.validate_on_submit():
        try:
            nome_banco = f"empresa_{form.cnpj.data.replace('.', '').replace('/', '').replace('-', '')}.db"
            clientes_dir = app.config.get('CLIENTES_DIR', 'clientes')
            os.makedirs(clientes_dir, exist_ok=True)
            caminho_banco = os.path.join(clientes_dir, nome_banco)
            empresa = Empresa(
                cnpj=form.cnpj.data,
                nome=form.nome.data,
                caminho_banco=caminho_banco
            )
            open(caminho_banco, 'a').close()
            db.session.add(empresa)
            db.session.commit()
            flash('Empresa cadastrada com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar empresa: {str(e)}', 'danger')
    else:
        flash('Dados inválidos para cadastro de empresa.', 'danger')
    return redirect(url_for('empresas'))

@app.route('/empresas/selecionar/<int:id>', methods=['POST'])
@login_required
def selecionar_empresa(id):
    empresa = Empresa.query.get_or_404(id)
    session['empresa_selecionada'] = empresa.id
    flash(f'Empresa selecionada: {empresa.nome}', 'info')
    return redirect(url_for('empresas'))

@app.route('/empresas/editar/<int:id>', methods=['POST'])
@login_required
def editar_empresa(id):
    try:
        empresa = Empresa.query.get_or_404(id)
        
        nome = request.form.get('nome')
        if nome:
            empresa.nome = nome
            db.session.commit()
            return jsonify({'status': 'success', 'message': 'Empresa atualizada com sucesso!'})
        else:
            return jsonify({'status': 'error', 'message': 'Nome da empresa é obrigatório'})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'Erro ao atualizar empresa: {str(e)}'})
    
@app.route('/empresas/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_empresa(id):
    try:
        empresa = Empresa.query.get_or_404(id)
        
        # Verificar se pode ser excluída ou deve ser inativada
        if empresa.pode_ser_excluida():
            # Pode excluir fisicamente
            db.session.delete(empresa)
            db.session.commit()
            flash('Empresa excluída com sucesso!', 'success')
        else:
            # Deve inativar
            empresa.inativar(current_user.username)
            db.session.commit()
            flash('Empresa inativada com sucesso! (Possui dados vinculados)', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao processar empresa: {str(e)}', 'error')
    
    return redirect(url_for('empresas'))

@app.route('/empresas/reativar/<int:id>', methods=['POST'])
@login_required
def reativar_empresa(id):
    try:
        empresa = Empresa.query.get_or_404(id)
        empresa.ativa = True
        empresa.data_inativacao = None
        empresa.usuario_inativacao = None
        db.session.commit()
        flash('Empresa reativada com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao reativar empresa: {str(e)}', 'error')
    
    return redirect(url_for('empresas'))


@app.route('/empresas/exportar/<int:id>')
@login_required
def exportar_banco(id):
    empresa = Empresa.query.get_or_404(id)
    return send_file(
        empresa.caminho_banco,
        as_attachment=True,
        download_name=f"backup_{empresa.cnpj}.db"
    )

@app.route('/empresas/importar', methods=['POST'])
@login_required
def importar_banco():
    arquivo = request.files.get('arquivo')
    if not arquivo or not arquivo.filename.endswith('.db'):
        flash('Selecione um arquivo .db válido.', 'danger')
        return redirect(url_for('empresas'))
    
    clientes_dir = app.config.get('CLIENTES_DIR', 'clientes')
    os.makedirs(clientes_dir, exist_ok=True)
    caminho = os.path.join(clientes_dir, arquivo.filename)
    arquivo.save(caminho)
    
    empresa = Empresa(
        cnpj=arquivo.filename.split('_')[1].replace('.db', ''),
        nome="Empresa Importada",
        caminho_banco=caminho
    )
    db.session.add(empresa)
    db.session.commit()
    flash('Banco de dados importado com sucesso!', 'success')
    return redirect(url_for('empresas'))

def criar_usuario_admin():
    """Cria usuário administrador padrão se não existir"""
    admin = Usuario.query.filter_by(username='admin').first()
    if not admin:
        admin = Usuario(
            username='admin',
            email='admin@elevalucro.com',
            nome_completo='Administrador ELEVALUCRO',
            is_admin=True,
            ativo=True
        )
        admin.set_password('123456')
        db.session.add(admin)
        db.session.commit()
        print("Usuário administrador criado: admin / 123456")

# Criar tabelas e usuário admin
with app.app_context():
    db.create_all()
    criar_usuario_admin()
    print("Banco de dados criado/verificado com sucesso!")

if __name__ == '__main__':
    print("Iniciando servidor Flask...")
    print("Acesse: http://127.0.0.1:5000")
    print("Login: admin / 123456")
    app.run(debug=True, host='127.0.0.1', port=5000)
